# '''
#     pymysql:
#     xlrd:读取excel表格
#     xlwt：写入excel表格
#     （.xls   .xlsx）
#     1.python -m pip install xlrd==0.93
#
#     任务1：
#         2020销售总额，所有数据统计出来
#     任务2：
#     excel <--> mysql
#
#
# '''
import xlrd

# import pymysql


wb = xlrd.open_workbook(filename=r"D:\day08\2020年每个月的销售情况.xlsx", encoding_override=True)

# data = sheet.cell_value(0,0)
# print(data)

# rows = sheet.nrows   # 获取多少行
# cols = sheet.ncols   # 获取多少列
# print(rows,"行",cols,"列")

# for i in range(0,rows):     # 获取多少行的数据
#     date = sheet.row_values(i)
#     print(date)
# for i in range(0,cols):     # 获取多少列的数据
#     date1 = sheet.col_values(i)
#     print(date1)


#
# 任务一：
# 全年的销售总额
# 所有工作表的名字列表
# ["1月","2月","3月","4月","5月","6月","7月","8月","9月","10月","11月","12月"]
# names = wb.sheet_names()
# total = 0
# # 循环每个工作表
# for i in names:
#     table = wb.sheet_by_name(i)
#     # 每个工作表的销售量列
#     count = table.col_values(-1)
#     # 每个工作表的单价列
#     price = table.col_values(-3)
#     sum1 = 0
#     for j in range(1, table.nrows):
#         sum1 += count[j] * price[j]
#     total += sum1
# print("全年销售总额: ", total)
#
# # 每件衣服的销售（件数）占比
# names = wb.sheet_names()
# sum = 0
# for i in names:
#     sheet = wb.sheet_by_name(i)
#     count = sheet.col_values(-1)
#     sum1 = 0
#     for j in range(1,sheet.nrows):
#         sum = sum + count[j]
#     sum1 += sum
# print("总销量：", sum1)
# set1 = set()
# for j in names:
#     sheet = wb.sheet_by_name(j)
#     count = sheet.col_values(1)
#     count = count[1:]
#     a = set(count)
#     set1 = set1 | a
#
# for x in set1:
#     print(x)
#     sum = 0
#     for a in names:
#         sheet = wb.sheet_by_name(a)
#         count = sheet.col_values(1)[1:]
#         count1 = sheet.col_values(-1)[1:]
#         row = sheet.nrows
#         for b in range(row-1):    #
#             if x == count[b]:
#                 sum += count1[b]
#     sum2 = sum / sum1
#     print("每件衣服的销售（件数）占比: %.2f"%sum2)

# 每件衣服的月销售占比
# names = wb.sheet_names()
# for i in names:
#     sum = 0
#     sheet = wb.sheet_by_name(i)
#     count = sheet.col_values(-1)
#     for j in range(1, sheet.nrows):
#         sum += count[j]
#
# set1 = set()
# for j in names:
#     sheet = wb.sheet_by_name(j)
#     count = sheet.col_values(1)
#     count = count[1:]
#     a = set(count)
#     set1 = set1 | a
#
# for x in set1:
#     print(x)
#     sum1 = 0
#     for a in names:
#         sheet = wb.sheet_by_name(a)
#         count = sheet.col_values(1)[1:]
#         count1 = sheet.col_values(-1)[1:]
#         row = sheet.nrows
#         for b in range(row - 1):
#                 if x == count[b]:
#                     sum1 += count1[b]
#         sum2 = sum1 / sum
#         print(f"{x}的{a}销售占比: %.2f"%sum2)


# 每件衣服的销售额占比
# names = wb.sheet_names()
# total = 0
# for i in names:
#     table = wb.sheet_by_name(i)
#     # 每个工作表的销售量列
#     count = table.col_values(-1)
#     # 每个工作表的单价列
#     price = table.col_values(-3)
#     sum1 = 0
#     for j in range(1, table.nrows):
#         sum1 += count[j] * price[j]
#     total += sum1
# print("全年销售总额: %.2f"%total)
#
# set1 = set()
# for j in names:
#     sheet = wb.sheet_by_name(j)
#     count = sheet.col_values(1)
#     count = count[1:]
#     a = set(count)
#     set1 = set1 | a
#
# for x in set1:
#     sum = 0
#     sum1 = 0
#     for a in names:
#         sheet = wb.sheet_by_name(a)
#         count = sheet.col_values(1)
#         count1 = sheet.col_values(2)
#         count2 = sheet.col_values(-1)
#         row = sheet.nrows
#         for b in range(1,row):
#             if x == count[b]:
#                 sum1 += count1[b] * count2[b]
#         sum2 =sum1 /total
#         print(f"{x}的销售额占比:%.2f"%(sum2*100),"%")



# 最畅销的衣服是那种   # 全年销量最低的衣服
# name = wb.sheet_names()
# set1 = set()
# list = []
# for j in name:
#     sheet = wb.sheet_by_name(j)
#     count = sheet.col_values(1)
#     count = count[1:]
#     a = set(count)
#     set1 = set1 | a
#
# for x in set1:
#     sum = 0
#     for i in name:
#         sheet = wb.sheet_by_name(i)
#         count = sheet.col_values(1)
#         count1 = sheet.col_values(2)
#         count2 = sheet.col_values(-1)
#         row = sheet.nrows
#         for j in range(1,row):
#             if x == count[j]:
#                 sum += count2[j]
#     list.append([x,sum])
#     a = min(list)
#     b = max(list)
# print("最畅销的衣服是：",a,"销量最低的衣服是:",b)



# 每个季度最畅销的衣服
# y=[[2,3,4],[5,6,7],[8,9,10],[11,0,1]]
# jd=0
# for x in range(0,4):
#     jd=jd+1
#     xlsum = 0
#     l = []
#     d = []
#     ll = []
#     for i in y[x]:
#         sheet = wb.sheet_by_index(i)
#         rows = sheet.nrows  # 获取多少行
#         cols = sheet.ncols  # 获取多少列
#         sum1 = 0
#         sum2 = 0
#         for j in range(1, rows):
#             data = sheet.row_values(j)
#             sum1 = sum1 + data[4]
#             if data[1] in l:
#                 continue
#             else:
#                 l.append(data[1])
#         xlsum = xlsum + sum1
#     for k in range(0, len(l)):
#         n = 0
#         m = 0
#         for i in y[x]:
#             sheet = wb.sheet_by_index(i)
#             rows = sheet.nrows  # 获取多少行
#             cols = sheet.ncols  # 获取多少列
#             for j in range(1, rows):
#                 data = sheet.row_values(j)
#                 if data[1] == l[k]:
#                     n = n + data[4]
#                     m = m + data[4] * data[2]
#         ll.append(n)
#         d.append((l[k],n))
#     temp=0
#     for j in range(len(ll)-1):
#         for i in range(len(ll)-j-1):
#             if ll[i]>ll[i+1]:
#                 temp=ll[i]
#                 ll[i]=ll[i+1]
#                 ll[i+1]=temp
#     for i in range(0, len(d)):
#         if d[i][1]==ll[0]:
#            print("第",jd,"季度最畅销的衣服是",d[i][0])






# 任务2：    excel <--> mysql
# 打开数据所在的工作簿
# book = xlrd.open_workbook(r'D:\day08\2020年每个月的销售情况.xlsx')
# # 选择存有数据的工作表
# sheet = book.sheet_by_name("12月")
#
# # 连接数据库
# con = pymysql.connect(host="localhost",user="root",password="",database="clothes_sell")
#
# cursor = con.cursor()  # 控制台
#
# # 创建插入SQL语句
# query = 'insert into 12月销售 (日期,服装名称,`价格/件`,本月库存数量,`销售量/每日`) values (%s, %s, %s, %s, %s)'
# # for循环迭代读取xls文件中的每行数据, 从第二行开始因为需要跳过标题行
# # 注意：xlrd方式表格的列标和行标都是从0开始计算
# for r in range(1, sheet.nrows):
#     num = sheet.cell(r, 0).value
#     name = sheet.cell(r, 1).value
#     age = sheet.cell(r, 2).value
#     classes = sheet.cell(r, 3).value
#     score = sheet.cell(r, 4).value
#     values = (num, name, age, classes, score)
#     # 执行sql语句
#     cursor.execute(query, values)
#
# # 数据库提交到数据库里
# con.commit()
#
# # 关闭资源
# cursor.close()
# con.close()
#
# # ncols和nrows表示获取excel表格的有效列数与行数
# columns = str(sheet.ncols)
# rows = str(sheet.nrows - 1)
# # 打印输出信息
# print("成功导入 " + columns + " 列 " + rows + " 行数据到MySQL数据库!")


# 任务三    MySQL    <-->    excel
import xlwt
wb = xlwt.Workbook()

st = wb.add_sheet("我家圆儿超可爱")

st.write(0,0,"部门编号")
st.write(0,1,"部门名称")


wb.save("三国集团.xls")


import openpyxl
import pymysql


# 连接数据库
con = pymysql.connect(host="localhost",user="root",password="",database="clothes_sell")

cursor = con.cursor()  # 控制台

wb = openpyxl.load_workbook('D:\day08\ceshi.xlsx')  # 创建一个工作簿
ws = wb.active

"""  获取表结构，并将表头写入excel """
cursor.execute("show tables;")
table_head = cursor.fetchall()
li = []

""" 获取表中所有数据，并写入excel """
cursor.execute("select * from 2月销售;")
table_product_data = cursor.fetchall()
j = 0
for da in table_product_data:
    di = 65
    j = j + 1
    for k in range(len(da)):

        ws[chr(di) + str(j)] = da[k]
        di = di+1

wb.save('D:\day08\ceshi.xlsx')

# # 关闭资源
cursor.close()
con.close()
